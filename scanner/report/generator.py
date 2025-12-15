# -*- coding: utf-8 -*-
"""
报告生成器：Excel/JSON/CSV。

聚焦排序（风险降序+规则 ID 升序）、字段完整性、配色与覆盖统计，面向 Python 3。
"""
import csv
import io
import json
import os

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover
    openpyxl = None
    PatternFill = None

SEVERITY_ORDER = ("high", "medium", "low")
SEVERITY_COLORS = {
    "high": "FFFF6B6B",    # 红
    "medium": "FFFFE066",  # 黄
    "low": "FFB2F2BB",     # 绿
}

HEADERS = ["rule_id", "group", "severity", "needs_review", "file", "line", "evidence", "reason", "suggestion"]

def _to_text(value):
    """确保输出为文本字符串。"""
    if value is None:
        return value
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except Exception:
            return value.decode("utf-8", "ignore")
    return value


def _severity_rank(severity):
    sev = (severity or "").lower()
    return SEVERITY_ORDER.index(sev) if sev in SEVERITY_ORDER else len(SEVERITY_ORDER)


def sort_findings(findings):
    """按风险降序、规则 ID 升序排序。"""
    return sorted(findings, key=lambda f: (_severity_rank(f.get("severity")), f.get("rule_id", "")))


def format_evidence(finding):
    """格式化证据，包含路径/行号/片段。"""
    file_path = _to_text(finding.get("file") or "-")
    line = finding.get("line")
    line_str = line if line is not None else "-"
    snippet = _to_text(finding.get("evidence") or "")
    parts = [u"路径:%s:%s" % (file_path, line_str)]
    if snippet:
        parts.append(u"片段:%s" % snippet)
    return u" | ".join(parts)


def _normalize_findings(findings):
    """补全字段并格式化证据/建议，保持字段顺序。"""
    normalized = []
    for f in sort_findings(findings):
        item = {}
        for key in HEADERS:
            value = f.get(key)
            if key in ("evidence", "reason", "suggestion", "rule_id", "group", "severity", "file"):
                item[key] = _to_text(value)
            else:
                item[key] = value
        item["evidence"] = format_evidence(f)
        item["reason"] = _to_text(f.get("reason") or "")
        item["suggestion"] = _to_text(f.get("suggestion") or "")
        normalized.append(item)
    return normalized


def generate_json_report(findings, output_path="report.json"):
    """生成 JSON 报告并返回路径。"""
    normalized_findings = _normalize_findings(findings)
    content = json.dumps(normalized_findings, ensure_ascii=False, indent=2)
    with io.open(output_path, "w", encoding="utf-8") as f:
        f.write(content)
    return output_path


def generate_csv_report(findings, output_path="report.csv"):
    """生成 CSV 报告并返回路径。"""
    normalized_findings = _normalize_findings(findings)
    with io.open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for item in normalized_findings:
            writer.writerow({key: item.get(key) for key in HEADERS})
    return output_path


def _build_coverage(findings):
    """按分组计算高/中/低/人工复核/总计，并追加合计。"""
    coverage = {}
    total = {"high": 0, "medium": 0, "low": 0, "needs_review": 0, "total": 0}
    for f in findings:
        group = _to_text(f.get("group") or "未分组")
        sev = (f.get("severity") or "").lower()
        needs_review = bool(f.get("needs_review"))

        if group not in coverage:
            coverage[group] = {"high": 0, "medium": 0, "low": 0, "needs_review": 0, "total": 0}

        if sev in ("high", "medium", "low"):
            coverage[group][sev] += 1
            total[sev] += 1
        coverage[group]["total"] += 1
        total["total"] += 1

        if needs_review:
            coverage[group]["needs_review"] += 1
            total["needs_review"] += 1

    coverage[_to_text("合计")] = total
    return coverage


def generate_excel_report(findings, output_path="report.xlsx"):
    """生成 Excel 报告，包含 Findings 与 Coverage 两个 Sheet。"""
    if openpyxl is None or PatternFill is None:  # pragma: no cover
        raise ImportError("缺少 openpyxl 依赖，无法生成 Excel 报告。")

    normalized_findings = _normalize_findings(findings)
    wb = openpyxl.Workbook()

    # Findings Sheet
    ws_findings = wb.active
    ws_findings.title = "Findings"
    ws_findings.append(HEADERS)
    for item in normalized_findings:
        row_values = [item.get(key) for key in HEADERS]
        ws_findings.append(row_values)

    # 配色：按 severity 列
    severity_col_index = HEADERS.index("severity") + 1  # 1-based
    for row in ws_findings.iter_rows(min_row=2, min_col=severity_col_index, max_col=severity_col_index):
        cell = row[0]
        sev = (cell.value or "").lower()
        color = SEVERITY_COLORS.get(sev)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Coverage Sheet
    ws_cov = wb.create_sheet("Coverage")
    ws_cov.append(["分组", "高", "中", "低", "人工复核", "总计"])

    coverage = _build_coverage(normalized_findings)
    total_key = _to_text("合计")
    for group in sorted(k for k in coverage.keys() if k != total_key):
        stats = coverage[group]
        ws_cov.append([group, stats["high"], stats["medium"], stats["low"], stats["needs_review"], stats["total"]])
    # 合计置底
    total_stats = coverage.get(total_key, {})
    if total_stats:
        ws_cov.append([total_key, total_stats.get("high", 0), total_stats.get("medium", 0), total_stats.get("low", 0), total_stats.get("needs_review", 0), total_stats.get("total", 0)])

    # 确保输出目录存在
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)
    wb.save(output_path)
    return output_path


__all__ = [
    "SEVERITY_COLORS",
    "HEADERS",
    "sort_findings",
    "format_evidence",
    "generate_json_report",
    "generate_csv_report",
    "generate_excel_report",
]
