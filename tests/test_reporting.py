# -*- coding: utf-8 -*-
"""
Reporting tests（TDD 占位）。

覆盖场景：
- Excel 字段/排序/配色/覆盖统计
- JSON/CSV 输出正确
"""

import csv
import json
import os
import tempfile
import unittest

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ReportingTests(unittest.TestCase):
    def setUp(self):
        self.findings = [
            {
                "rule_id": "PAY-002",
                "group": "PAY",
                "severity": "high",
                "file": "app/pay.m",
                "line": 10,
                "evidence": "调用第三方支付 SDK",
                "reason": "外链或第三方支付",
                "suggestion": "数字内容/服务需使用 IAP，移除外链或改为 IAP。",
                "needs_review": True,
            },
            {
                "rule_id": "PRIV-001",
                "group": "PRIV",
                "severity": "high",
                "file": "app/privacy.m",
                "line": 25,
                "evidence": "缺少权限用途文案",
                "reason": "权限文案缺失",
                "suggestion": "为权限添加清晰的用途说明。",
                "needs_review": False,
            },
            {
                "rule_id": "AUTH-003",
                "group": "AUTH",
                "severity": "medium",
                "file": "app/login.swift",
                "line": 15,
                "evidence": "第三方登录实现",
                "reason": "缺少 Sign in with Apple",
                "suggestion": "在同等位置提供 Sign in with Apple。",
                "needs_review": True,
            },
            {
                "rule_id": "NET-HTTP",
                "group": "NET",
                "severity": "low",
                "file": "app/network.swift",
                "line": 88,
                "evidence": "出现 http://example.com",
                "reason": "明文 HTTP",
                "suggestion": "改用 HTTPS 或配置 ATS 例外并说明。",
                "needs_review": True,
            },
        ]
        self.expected_headers = [
            "rule_id",
            "group",
            "severity",
            "needs_review",
            "file",
            "line",
            "evidence",
            "reason",
            "suggestion",
        ]
        self.expected_order = ["PAY-002", "PRIV-001", "AUTH-003", "NET-HTTP"]
        self.coverage_expected = {
            "PAY": {"high": 1, "medium": 0, "low": 0, "needs_review": 1, "total": 1},
            "PRIV": {"high": 1, "medium": 0, "low": 0, "needs_review": 0, "total": 1},
            "AUTH": {"high": 0, "medium": 1, "low": 0, "needs_review": 1, "total": 1},
            "NET": {"high": 0, "medium": 0, "low": 1, "needs_review": 1, "total": 1},
            "合计": {"high": 2, "medium": 1, "low": 1, "needs_review": 3, "total": 4},
        }
        self.severity_order = ("high", "medium", "low")

    def test_excel_output(self):
        """验证字段完整、配色规则、风险降序+规则ID 升序、覆盖统计数值。"""
        try:
            from scanner.report.generator import SEVERITY_COLORS, generate_excel_report
        except ImportError:
            self.skipTest("报告生成器未实现，暂跳过验证。")
            return
        if openpyxl is None:
            self.skipTest("缺少 openpyxl 依赖，暂跳过验证。")
            return

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = os.path.join(tmp_dir, "report.xlsx")
            result_path = generate_excel_report(self.findings, output_path=output_path)

            self.assertTrue(os.path.exists(result_path), "Excel 报告应生成到指定路径。")

            workbook = openpyxl.load_workbook(result_path)
            self.assertIn("Findings", workbook.sheetnames, "应包含 Findings 明细 Sheet。")
            self.assertIn("Coverage", workbook.sheetnames, "应包含 Coverage 覆盖统计 Sheet。")

            findings_sheet = workbook["Findings"]
            headers = [cell.value for cell in findings_sheet[1]]
            self.assertEqual(headers, self.expected_headers, "字段顺序或命名不一致。")

            # 校验排序：风险降序（高>中>低），同等级按规则 ID 升序。
            rows = list(findings_sheet.iter_rows(min_row=2, values_only=True))
            rule_ids = [row[0] for row in rows]
            self.assertEqual(rule_ids, self.expected_order, "排序应按风险降序、规则ID 升序。")

            severity_index = headers.index("severity")
            findings_by_rule = {f["rule_id"]: f for f in self.findings}

            for row_idx, row in enumerate(findings_sheet.iter_rows(min_row=2, values_only=False)):
                values = [cell.value for cell in row]
                rule_id = values[0]
                expected = findings_by_rule[rule_id]
                for idx, header in enumerate(self.expected_headers):
                    self.assertEqual(values[idx], expected[header], "字段 %s 填充不正确。" % header)

                severity_value = values[severity_index]
                self.assertIn(severity_value, self.severity_order, "风险等级应为 high/medium/low。")
                fill_rgb = row[severity_index].fill.start_color.rgb
                expected_rgb = SEVERITY_COLORS.get(severity_value)
                self.assertEqual(fill_rgb, expected_rgb, "配色应符合高红/中黄/低绿规则。")

            # 覆盖统计 Sheet：按分组统计高/中/低/人工复核/总计。
            coverage_sheet = workbook["Coverage"]
            coverage_headers = [cell.value for cell in coverage_sheet[1]]
            self.assertEqual(
                coverage_headers,
                ["分组", "高", "中", "低", "人工复核", "总计"],
                "覆盖统计表头不符合约定。",
            )

            for row in coverage_sheet.iter_rows(min_row=2, values_only=True):
                group = row[0]
                if group not in self.coverage_expected:
                    continue
                expected_stats = self.coverage_expected[group]
                self.assertEqual(row[1], expected_stats["high"], "%s 高风险数量不匹配。" % group)
                self.assertEqual(row[2], expected_stats["medium"], "%s 中风险数量不匹配。" % group)
                self.assertEqual(row[3], expected_stats["low"], "%s 低风险数量不匹配。" % group)
                self.assertEqual(row[4], expected_stats["needs_review"], "%s 人工复核数量不匹配。" % group)
                self.assertEqual(row[5], expected_stats["total"], "%s 总计数量不匹配。" % group)

    def test_json_csv_output(self):
        """验证 JSON/CSV 输出正确。"""
        try:
            from scanner.report.generator import generate_csv_report, generate_json_report
        except ImportError:
            self.skipTest("报告生成器未实现，暂跳过 JSON/CSV 验证。")
            return

        with tempfile.TemporaryDirectory() as tmp_dir:
            json_path = os.path.join(tmp_dir, "report.json")
            csv_path = os.path.join(tmp_dir, "report.csv")

            json_result = generate_json_report(self.findings, output_path=json_path)
            csv_result = generate_csv_report(self.findings, output_path=csv_path)

            self.assertTrue(os.path.exists(json_result), "JSON 报告应生成到指定路径。")
            self.assertTrue(os.path.exists(csv_result), "CSV 报告应生成到指定路径。")

            with open(json_result, "r") as jf:
                data = json.load(jf)
            self.assertIsInstance(data, list, "JSON 输出应为列表。")
            rule_ids = [item.get("rule_id") for item in data]
            self.assertEqual(rule_ids, self.expected_order, "JSON 输出排序应按风险降序、规则ID 升序。")
            for item in data:
                for header in self.expected_headers:
                    self.assertIn(header, item, "JSON 缺少字段 %s。" % header)

            with open(csv_result, "r") as cf:
                reader = csv.DictReader(cf)
                self.assertEqual(reader.fieldnames, self.expected_headers, "CSV 字段顺序不一致。")
                rows = list(reader)
            csv_rule_ids = [row["rule_id"] for row in rows]
            self.assertEqual(csv_rule_ids, self.expected_order, "CSV 输出排序应按风险降序、规则ID 升序。")
            # CSV 将数值转为字符串，做字符串对比
            for row in rows:
                source = next(f for f in self.findings if f["rule_id"] == row["rule_id"])
                for header in self.expected_headers:
                    self.assertEqual(str(row[header]), str(source[header]), "CSV 字段 %s 填充不正确。" % header)


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
